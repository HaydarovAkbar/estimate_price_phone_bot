from telegram import Update, ReplyKeyboardRemove
from telegram.ext import CallbackContext
import openpyxl
import time

from .keyboards import AdminKeyboards as K
from .message import MessageText as T

from states import States as S
from db.models import User, Channels, Categories, Capacities, Products, Documents, Countries, Statuses, Memories, \
    Colors, ProductCriteria, Prices


def admin(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    update.message.reply_text(T().main[user_lang].format(tg_user.full_name),
                              reply_markup=K().base(user_lang))
    return S.ADMIN


def add_admin(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    update.message.reply_text(T().add_admin[user_lang],
                              reply_markup=K().back(user_lang))
    return S.ADD_ADMIN


def get_admin_id(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    try:
        user_id = int(update.message.text)
    except ValueError:
        update.message.reply_text(T().error_id[user_lang])
        return S.ADD_ADMIN
    user, _ = User.objects.get_or_create(chat_id=user_id,
                                         defaults={'chat_id': user_id, 'language': user_lang, 'is_active': True,
                                                   'is_admin': True})
    if _ or not user.is_admin:
        user.is_admin = True
        user.save()
        update.message.reply_text(T().success[user_lang], reply_markup=K().base(user_lang))
    else:
        update.message.reply_text(T().already_admin[user_lang],
                                  reply_markup=K().base(user_lang))
    return S.ADMIN


def get_admins(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    admins = User.objects.filter(is_active=True, is_admin=True)
    if admins.exists():
        text = T().admins[user_lang]
        i = 1
        for admin in admins:
            text += f"{i}) {admin.get_fullname()}\n"
            i += 1
        # text += T().delete_admin[user_lang]
        update.message.reply_text(text, reply_markup=K().users(admins))
        update.message.reply_text(T().delete_admin[user_lang], reply_markup=K().back(user_lang))
        return S.ADMINS
    else:
        update.message.reply_text(T().no_admins[user_lang], reply_markup=K().base(user_lang))
    return S.ADMIN


def delete_admin(update: Update, context: CallbackContext):
    tg_user = update.callback_query.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    user_id = int(update.callback_query.data)
    user = User.objects.get(chat_id=user_id)
    user.is_admin = False
    user.save()
    update.callback_query.delete_message()
    update.callback_query.bot.send_message(chat_id=tg_user.id, text=T().success[user_lang],
                                           reply_markup=K().base(user_lang))
    return S.ADMIN


def get_users(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    all_text = "{:<1} {:<8} {:<23}\n\n".format("ID ||", "   Name      ||", "   Date")
    users = User.objects.filter(is_active=True).order_by('created_at')[0:50]
    i = 1
    for user in users[::-1]:
        all_text += "{:<1} | {:<8} | {:<23}\n".format(i, user.fullname,
                                                      user.created_at.strftime('%Y-%m-%d %H:%M'))
        i += 1
    all_user_count = f"Jami foydalanuvchilar soni: {User.objects.all().count()}\n\n"
    all_user_count += all_text
    update.message.reply_text(all_user_count, reply_markup=K().base(user_lang))
    return S.ADMIN


def add_data(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    update.message.reply_text(T().add_data[user_lang],
                              reply_markup=K().back(user_lang))
    update.message.reply_document(document=open('static/data.xlsx', 'rb'))
    return S.ADD_DATA


def change_data(path='static/data.xlsx'):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    Capacities.objects.all().delete()
    Categories.objects.all().delete()
    Documents.objects.all().delete()
    Countries.objects.all().delete()
    Statuses.objects.all().delete()
    Memories.objects.all().delete()
    Colors.objects.all().delete()
    Products.objects.all().delete()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if '#end' in row[0].value:
            break
        category, _ = Categories.objects.get_or_create(title=row[0].value.strip())
        capacity, _ = Capacities.objects.get_or_create(title=row[2].value.strip())
        document, _ = Documents.objects.get_or_create(title=row[5].value.strip())
        country, _ = Countries.objects.get_or_create(title=row[6].value.strip())
        status, _ = Statuses.objects.get_or_create(title=row[7].value.strip())
        memory, _ = Memories.objects.get_or_create(title=row[4].value.strip())
        color, _ = Colors.objects.get_or_create(title=row[3].value.strip())
        # price, _ = Prices.objects.get_or_create(title=row[8].value.strip())
        producty, _ = Products.objects.get_or_create(title=row[1].value.strip(), category=category)
        data = dict()
        if capacity.title != 'Y':
            data['capacity'] = capacity
        if color.title != 'Y':
            data['color'] = color
        if memory.title != 'Y':
            data['memory'] = memory
        if document.title != 'Y':
            data['document'] = document
        if country.title != 'Y':
            data['country'] = country
        if status.title != 'Y':
            data['status'] = status
        if row[8].value.strip() != 'Y':
            data['price'] = row[8].value.strip()
        ProductCriteria.objects.create(product=producty, **data)
    wb.save(path)


def get_data(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    file_id = update.message.document.file_id
    file_name = update.message.document.file_name
    file = context.bot.get_file(file_id)
    update.message.reply_text(T().wait[user_lang])
    file.download('static/' + file_name)
    change_data('static/' + file_name)
    update.message.reply_text(T().success[user_lang], reply_markup=K().base(user_lang))
    return S.ADMIN


def add_channel(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    update.message.reply_text(T().add_channel[user_lang],
                              reply_markup=K().back(user_lang))
    return S.ADD_CHANNEL


def get_channel_name(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    channel_name = update.message.text
    context.user_data['channel_name'] = channel_name
    update.message.reply_text(T().add_channel_url[user_lang],
                              reply_markup=K().back(user_lang))
    return S.ADD_CHANNEL_URL


def get_channel_url(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    channel_url = update.message.text
    context.user_data['channel_url'] = channel_url
    update.message.reply_text(T().add_channel_id[user_lang], reply_markup=K().back(user_lang))
    return S.ADD_CHANNEL_ID


def get_channel_id(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    channel_id = update.message.text
    channel, _ = Channels.objects.get_or_create(title=context.user_data['channel_name'],
                                                defaults={'title': context.user_data['channel_name'],
                                                          'chat_id': channel_id,
                                                          'url': context.user_data['channel_url'],
                                                          'is_active': True})
    if _:
        update.message.reply_text(T().channel_added[user_lang], reply_markup=K().base(user_lang))
    else:
        update.message.reply_text(T().channel_already_exists[user_lang],
                                  reply_markup=K().base(user_lang))
    return S.ADMIN


def get_channels(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    channels = Channels.objects.filter(is_active=True)
    if channels.exists():
        update.message.reply_text(T().channels[user_lang], reply_markup=K().all_channels(channels))
    else:
        update.message.reply_text(T().no_channels[user_lang], reply_markup=K().base(user_lang))
        return S.ADMIN
    return S.DELETE_CHANNEL


def delete_channel(update: Update, context: CallbackContext):
    tg_user = update.callback_query.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    channel_id = int(update.callback_query.data)
    try:
        Channels.objects.get(chat_id=channel_id).delete()
    except Exception:
        update.callback_query.delete_message()
        context.bot.send_message(chat_id=tg_user.id, text=T().error_channel[user_lang])
        return S.ADMIN
    update.callback_query.delete_message()
    update.callback_query.bot.send_message(chat_id=tg_user.id, text=T().channel_deleted[user_lang])
    return S.ADMIN


def get_reklama(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    user = user.first()
    user_lang = user.language if user.language else 'uz'
    update.message.reply_text(T().fill_template[user_lang], reply_markup=K().back(user_lang))
    return S.REKLAMA


def send_reklama(update: Update, context: CallbackContext):
    tg_user = update.message.from_user
    user = User.objects.filter(chat_id=tg_user.id, is_active=True, is_admin=True)
    if not user.exists():
        return S.START
    counter = 0
    count_user = User.objects.all().count()
    for other_user in User.objects.all():
        if counter % 5000 == 0 and counter != 0:
            for admin in User.objects.filter(is_admin=True):
                try:
                    context.bot.send_message(chat_id=admin.chat_id,
                                             text=f"Reklama!\n{count_user} - foydalanuvchidan\n{counter} - foydalanuvchiga foydalanuvchiga yuborildi")
                except Exception:
                    pass
        try:
            update.message.copy(chat_id=other_user.chat_id)
            counter += 1
            time.sleep(20)
        except Exception:
            pass
    update.message.reply_text(f"✅ Reklama {counter} ta foydalanuvchiga yuborildi ✅",
                              reply_markup=K().base('uz'))
    return S.ADMIN
